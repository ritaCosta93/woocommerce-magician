import random
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Generate random data for products
categories = ["Electronics", "Clothing", "Books", "Home & Kitchen", "Toys", "Sports"]
subcategories = {
    "Electronics": ["Smartphones", "Laptops", "Tablets", "Accessories"],
    "Clothing": ["Men's", "Women's", "Kids'"],
    "Books": ["Fiction", "Non-Fiction", "Children's"],
    "Home & Kitchen": ["Furniture", "Appliances", "Cookware"],
    "Toys": ["Action Figures", "Puzzles", "Board Games"],
    "Sports": ["Fitness", "Outdoor", "Team Sports"]
}

brands = ["Brand1", "Brand2", "Brand3", "Brand4", "Brand5", "Brand6"]
site_statuses = ["published", "draft"]

# Generate random products
products = []
for _ in range(20):
    category = random.choice(categories)
    subcategory = random.choice(subcategories[category])
    product_name = f"{category} {subcategory} Product {random.randint(1, 100)}"
    description = f"Description for {product_name}"
    price = round(random.uniform(10, 1000), 2)
    ref = f"REF{random.randint(1000, 9999)}"
    brand = random.choice(brands)
    site_status = random.choice(site_statuses)
    image_url = f"https://example.com/images/{ref}.jpg"  # Placeholder URL for image

    products.append({
        "Category": category,
        "Subcategory": subcategory,
        "Product Name": product_name,
        "Description": description,
        "Price": price,
        "Reference (SKU)": ref,
        "Brand": brand,
        "Site Status": site_status,
        "Image URL": image_url  # New column for image URL
    })

# Create an Excel workbook and sheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Products"

# Write headers
headers = ["Category", "Subcategory", "Product Name", "Description", "Price", "Reference (SKU)", "Brand", "Site Status", "Image URL"]
for col_idx, header in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col_idx)
    cell.value = header
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Write data
for row_idx, product in enumerate(products, start=2):
    for col_idx, key in enumerate(headers, start=1):
        cell = sheet.cell(row=row_idx, column=col_idx)
        cell.value = product[key]
        cell.alignment = Alignment(horizontal='left', vertical='center')

# Save workbook
file_path = "products.xlsx"
wb.save(file_path)

print(f"Random products sheet generated successfully: {file_path}")
