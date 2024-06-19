import os
import json
import asyncio
import logging
import subprocess
import sys
from typing import List, Dict, Set, Tuple
from pathlib import Path

# Ensure required libraries are installed
def install_requirements():
    requirements_file = 'requirements.txt'
    if not os.path.exists(requirements_file):
        with open(requirements_file, 'w') as f:
            f.write("openpyxl\nrequests\nratelimit\nwoocommerce\n")
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', '-r', requirements_file])

install_requirements()

import openpyxl
from woocommerce import API
from ratelimit import limits, sleep_and_retry

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class WooCommerceService:
    def __init__(self, config):
        self.config = config
        self.woocommerce = API(
            url=config['WOOCOMMERCE_URL'],
            consumer_key=config['WOOCOMMERCE_CONSUMER_KEY'],
            consumer_secret=config['WOOCOMMERCE_CONSUMER_SECRET'],
            version="wc/v3",
            timeout=5  # Increased timeout for slower network conditions
        )

    def handle_cron(self):
        logger.info("Scheduled task triggered.")
        asyncio.run(self.upload())

    async def upload(self):
        logger.info("Starting upload process...")
        await self.upload_products_from_excel()

    async def upload_products_from_excel(self):
        local_file_path = 'products.xlsx'
        if not os.path.exists(local_file_path):
            logger.error(f"Local file not found: {local_file_path}")
            return
        
        logger.info(f"Reading Excel file from: {local_file_path}")
        workbook = openpyxl.load_workbook(local_file_path)
        worksheet = workbook.active

        categories = set()
        subcategories = {}

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            category_name = str(row[0]).strip() if row[0] else ''
            subcategory_name = str(row[1]).strip() if row[1] else ''
            if category_name:
                categories.add(category_name)
            if subcategory_name and category_name:
                subcategories[subcategory_name] = category_name

        category_map = await self.create_categories_and_subcategories(categories, subcategories)
        existing_products = await self.fetch_all_products()
        existing_images = await self.fetch_all_media()
        updated_products = []

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            try:
                image_path = self.normalize_image_path(row[2])
                if os.path.exists(image_path):
                    image_id, image_url = await self.find_or_upload_image(existing_images, image_path)
                else:
                    image_id, image_url = None, None

                product_data = await self.prepare_product_data(row, category_map, image_id, image_url)

                if product_data['categories']:
                    existing_product = next((p for p in existing_products if p['sku'] == product_data['sku']), None)
                    if existing_product:
                        await self.update_product_with_semaphore(existing_product['id'], product_data)
                    else:
                        await self.create_product_with_semaphore(product_data)
                    updated_products.append(product_data)
                else:
                    logger.error(f"Product {product_data['name']} not created: missing categories")

            except Exception as e:
                logger.error(f"Error processing product row: {row} - {str(e)}")

        logger.info("Upload process completed.")
        self.generate_json_file(updated_products)
        await self.update_excel_file_with_image_urls(local_file_path, updated_products)

    async def create_categories_and_subcategories(self, categories: Set[str], subcategories: Dict[str, str]) -> Dict[str, int]:
        category_map = {}

        for category_name in categories:
            category_id = await self.ensure_category_exists(category_name)
            logger.info(f"Category created/exists: {category_name} with ID {category_id}")
            category_map[category_name] = category_id

        for subcategory_name, parent_category_name in subcategories.items():
            parent_category_id = category_map.get(parent_category_name)
            if parent_category_id:
                subcategory_id = await self.ensure_category_exists(subcategory_name, parent_category_id)
                logger.info(f"Subcategory created/exists: {subcategory_name} with ID {subcategory_id}, parent ID {parent_category_id}")
                category_map[subcategory_name] = subcategory_id
            else:
                logger.error(f"Parent category {parent_category_name} not found for subcategory {subcategory_name}")

        return category_map

    async def fetch_all_products(self) -> List[Dict]:
        logger.info("Fetching all products from WooCommerce...")
        products = []
        page = 1
        per_page = 100

        while True:
            response = await asyncio.to_thread(
                self.woocommerce.get, "products", params={'page': page, 'per_page': per_page}
            )

            if response.status_code == 200:
                products_data = response.json()
                if not products_data:
                    break
                products.extend(products_data)
                page += 1
            else:
                logger.error(f"Failed to fetch products. Status code: {response.status_code}, Response: {response.content}")
                response.raise_for_status()  # Raise an exception for non-successful status codes

        return products

    async def fetch_all_media(self) -> List[Dict]:
        logger.info("Fetching all media (images) from WooCommerce...")
        media = []
        page = 1
        per_page = 100

        try:
            while True:
                response = await asyncio.to_thread(
                    self.woocommerce.get, "media", params={'page': page, 'per_page': per_page}
                )

                response.raise_for_status()  # Raise an exception for non-successful status codes

                if response.status_code == 200:
                    media_data = response.json()
                    if not media_data:
                        break
                    media.extend(media_data)
                    page += 1
                else:
                    logger.error(f"Failed to fetch media. Status code: {response.status_code}, Response: {response.content}")
                    break  # Exit loop on error
        except Exception as e:
            logger.error(f"Error fetching media: {str(e)}")

        return media

    async def ensure_category_exists(self, name: str, parent_id: int = 0) -> int:
        try:
            response = await asyncio.to_thread(
                self.woocommerce.get, "products/categories", params={'search': name}
            )
            
            if response.status_code == 200:
                existing_categories = response.json()
                if isinstance(existing_categories, list) and existing_categories:
                    category_id = existing_categories[0]['id']
                    logger.info(f"Category {name} exists with ID {category_id}")
                    return category_id
                else:
                    new_category_response = await asyncio.to_thread(
                        self.woocommerce.post, "products/categories", data={'name': name, 'parent': parent_id}
                    )
                    new_category = new_category_response.json()
                    category_id = new_category['id']
                    logger.info(f"Category {name} created with ID {category_id}")
                    return category_id
            else:
                logger.error(f"Failed to fetch categories. Status code: {response.status_code}, Response: {response.content}")
                response.raise_for_status()  # Raise an exception for non-successful status codes

        except Exception as e:
            logger.error(f"Error ensuring category exists {name}: {str(e)}")
            return 0

    @sleep_and_retry
    @limits(calls=5, period=15)
    async def create_product_with_semaphore(self, product_data: Dict):
        try:
            logger.info(f"Creating product: {product_data['name']}")
            await asyncio.to_thread(self.woocommerce.post, "products", data=product_data)
            logger.info(f"Created product: {product_data['name']}")
        except Exception as e:
            if "already exists" in str(e):
                logger.warning(f"Product {product_data['name']} already exists, skipping creation.")
            else:
                logger.error(f"Error creating product {product_data['name']}: {str(e)}")

    @sleep_and_retry
    @limits(calls=5, period=15)
    async def update_product_with_semaphore(self, product_id: int, product_data: Dict):
        try:
            logger.info(f"Updating product: {product_data['name']} (ID: {product_id})")
            await asyncio.to_thread(self.woocommerce.put, f"products/{product_id}", data=product_data)
            logger.info(f"Updated product: {product_data['name']} (ID: {product_id})")
        except Exception as e:
            if "already exists" in str(e):
                logger.warning(f"Product {product_data['name']} (ID: {product_id}) already up to date, skipping update.")
            else:
                logger.error(f"Error updating product {product_data['name']} (ID: {product_id}): {str(e)}")

    def generate_json_file(self, data):
        logger.info("Generating JSON file with updated products...")
        with open('updated_products.json', 'w') as json_file:
            json.dump(data, json_file, indent=4)
        logger.info("JSON file generated successfully.")

    async def update_excel_file_with_image_urls(self, local_file_path: str, updated_products: List[Dict]):
        logger.info("Updating Excel file with image URLs...")
        workbook = openpyxl.load_workbook(local_file_path)
        worksheet = workbook.active

        for row, product in zip(worksheet.iter_rows(min_row=2), updated_products):
            row[-1].value = product.get('imageUrl', '')

        workbook.save(local_file_path)
        logger.info("Excel file updated successfully.")

    def normalize_image_path(self, image_path: str) -> str:
        base_local_dir = os.getenv('IMAGES_PATH', 'images')
        return os.path.join(base_local_dir, image_path)

    async def find_or_upload_image(self, existing_images: List[Dict], image_path: str) -> Tuple[int, str]:
        for image in existing_images:
            if image['source_url'] == image_path:
                return image['id'], image['source_url']

        return await self.upload_image_to_woocommerce(image_path)

    @sleep_and_retry
    @limits(calls=5, period=15)
    async def upload_image_to_woocommerce(self, image_path: str) -> Tuple[int, str]:
        try:
            logger.info(f"Uploading image: {image_path}")
            with open(image_path, 'rb') as img_file:
                response = await asyncio.to_thread(
                    self.woocommerce.post, "media", files={'file': img_file}
                )
                image_data = response.json()
                image_id = image_data['id']
                image_url = image_data['source_url']
                logger.info(f"Uploaded image {image_path} to WooCommerce with ID {image_id}")
                return image_id, image_url
        except Exception as e:
            if "already exists" in str(e):
                logger.warning(f"Image {image_path} already exists, skipping upload.")
            else:
                logger.error(f"Error uploading image {image_path} to WooCommerce: {str(e)}")
            return None, ''

    async def prepare_product_data(self, row: Tuple, category_map: Dict[str, int], image_id: int = None, image_url: str = None) -> Dict:
        try:
            product_data = {
                'name': str(row[2]).strip() if row[2] else '',
                'type': 'simple',
                'regular_price': str(row[4]) if isinstance(row[4], (int, float)) else '0',
                'description': str(row[3]).strip() if row[3] else '',
                'short_description': str(row[3]).strip() if row[3] else '',
                'sku': str(row[5]).strip() if row[5] else '',
                'categories': []
            }

            if image_id and image_url:
                product_data['images'] = [{'id': image_id, 'src': image_url}]

            category_name = str(row[0]).strip() if row[0] else ''
            subcategory_name = str(row[1]).strip() if row[1] else ''

            if category_name:
                category_id = category_map.get(category_name)
                if category_id:
                    product_data['categories'].append({'id': category_id})

            if subcategory_name:
                subcategory_id = category_map.get(subcategory_name)
                if subcategory_id:
                    product_data['categories'].append({'id': subcategory_id})

            return product_data
        except Exception as e:
            logger.error(f"Error preparing product data: {str(e)}")
            raise

if __name__ == "__main__":
    config = {
        'WOOCOMMERCE_URL': '',
        'WOOCOMMERCE_CONSUMER_KEY': '',
        'WOOCOMMERCE_CONSUMER_SECRET': ''
    }

    service = WooCommerceService(config)
    service.handle_cron()
